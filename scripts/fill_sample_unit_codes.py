"""Batch 5 Part 1 — put a REAL official unit code in column J of the Excel
import samples.

Column J («کد واحد رسمی») is optional, and the shipped samples left it empty in
every data row. That was harmless while `tax_units` was empty, but now that the
official RC_UMGS.ST_V1.18 catalog is imported the importer validates J against
it — so the sample an accountant downloads should SHOW what a valid value looks
like instead of leaving them to guess.

Run inside the api container (it has openpyxl):

    docker run --rm -v "$PWD/../digi-tax-frontend:/fe" -v "$PWD:/ops" \
      digitax-backend:local python /ops/scripts/fill_sample_unit_codes.py /fe/public/samples

Idempotent: a row that already carries a code is left alone. The codes used here
MUST exist in digi-tax-backend/data/moadian/rc_umgs_st_v1_18_units.csv.
"""

from __future__ import annotations

import sys
from pathlib import Path

# عدد (1627) is the everyday default; کیلوگرم (164) and ساعت (16103) show that a
# goods row and a service row can legitimately differ. All three are verbatim
# RC_UMGS.ST_V1.18 rows — never invent a code.
UNIT_CYCLE = ["1627", "164", "16103"]
UNIT_COLUMN = 10  # J


def fill(path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    written = 0
    data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, 6).value not in (None, "")]
    for i, row in enumerate(data_rows):
        cell = ws.cell(row, UNIT_COLUMN)
        if cell.value in (None, ""):
            cell.value = UNIT_CYCLE[i % len(UNIT_CYCLE)]
            written += 1
    if written:
        wb.save(path)
    return written


def main(folder: str) -> int:
    base = Path(folder)
    files = sorted(base.glob("moadian-import-sample*.xlsx"))
    if not files:
        print(f"no sample workbooks under {base}", file=sys.stderr)
        return 2
    for f in files:
        print(f"{f.name}: wrote {fill(f)} unit codes")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1] if len(sys.argv) > 1 else "public/samples"))
