"""Accountant-batch micro-cleanup — add column Q «تعداد به واحد دوم» to the
Excel import samples.

Part 4 taught the importer an optional 17th column: on a row resolved by بارکد
(column P) whose product carries a «واحد دوم» + ضریب تبدیل, column Q lets the
merchant type the quantity in the PACKAGING unit (the stone-shop's ۹۵ جعبه)
and the importer computes the primary official quantity itself. The shipped
samples still ended at column P, so nobody downloading one could discover the
column exists. This adds the HEADER only — the data rows stay empty, because a
non-empty Q is an error unless the row's barcode resolves a dual-unit product,
and the sample rows deliberately carry no barcode.

Run from digi-tax-ops with the stack up (the api image has openpyxl):

    docker run --rm -v "$PWD/../digi-tax-frontend:/fe" -v "$PWD:/ops" \
      digi-tax-ops-api:latest python /ops/scripts/add_sample_second_unit_column.py /fe/public/samples

Idempotent: a sample whose Q header is already present is left untouched.
"""

from __future__ import annotations

import sys
from pathlib import Path

HEADER = "تعداد به واحد دوم (اختیاری — فقط با بارکد کالای دارای واحد دوم)"
COLUMN = 17  # Q


def add_header(path: Path) -> bool:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    cell = ws.cell(1, COLUMN)
    if cell.value not in (None, ""):
        return False
    template = ws.cell(1, COLUMN - 1)  # copy the P header's look
    cell.value = HEADER
    if template.has_style:
        cell.font = template.font.copy()
        cell.fill = template.fill.copy()
        cell.alignment = template.alignment.copy()
        cell.border = template.border.copy()
    ws.column_dimensions["Q"].width = 34
    wb.save(path)
    return True


def main() -> int:
    samples_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
    files = sorted(samples_dir.glob("moadian-import-sample*.xlsx"))
    if not files:
        print(f"no moadian-import-sample*.xlsx under {samples_dir}", file=sys.stderr)
        return 1
    for f in files:
        changed = add_header(f)
        print(f"{f.name}: {'Q header added' if changed else 'already present'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
